"""
Build G.E.T. Rental Trucks workbook: master data + customer insurance view.
Excel 365 / Excel on the web compatible (tables, dynamic arrays, no VBA).
"""
from __future__ import annotations

import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.properties import CalcProperties
from openpyxl.workbook.defined_name import DefinedName

SOURCE = Path(r"c:\Users\Jason Jones\Downloads\G.E.T. Rental Trucks.xlsx")
OUTPUT = Path(r"c:\Users\Jason Jones\Downloads\G.E.T. Rental Trucks - Workbook.xlsx")
BACKUP = Path(r"c:\Users\Jason Jones\Downloads\G.E.T. Rental Trucks - Original Backup.xlsx")

MASTER_SHEET = "Master - Rental Units"
CUSTOMER_SHEET = "Customer Insurance"
GUIDE_SHEET = "How to Use"
TABLE_NAME = "RentalMaster"
MAX_DATA_ROWS = 200
DATA_START_ROW = 5

# Colors
NAVY = "1F3864"
BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
LIGHT_GREEN = "E2EFDA"
LIGHT_YELLOW = "FFF2CC"
WHITE = "FFFFFF"
ALT_ROW = "F2F7FB"
WARN = "FFC7CE"
WARN_TEXT = "9C0006"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=11, color="404040")
BODY_FONT = Font(name="Calibri", size=11)
SMALL_FONT = Font(name="Calibri", size=9, color="666666")

HEADER_FILL = PatternFill("solid", fgColor=BLUE)
TITLE_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
ALT_FILL = PatternFill("solid", fgColor=ALT_ROW)
WARN_FILL = PatternFill("solid", fgColor=WARN)

THIN = Side(style="thin", color="B4C6E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MASTER_HEADERS = [
    "Stock #",
    "Make",
    "Year",
    "Asset Value ($)",
    "ON ROAD VIN",
    "Serial Number",
    "Model",
    "Meter Out",
    "Meter In",
    "Fuel Out",
    "Fuel In",
    "Date Out",
    "Date In",
    "Rental Status",
    "Needs VIN",
]

CUSTOMER_HEADERS = [
    "Unit Number",
    "VIN (17 characters)",
    "Serial Number",
    "Model Year",
    "Meter at Pickup",
    "Meter at Return",
    "Fuel at Pickup",
    "Fuel at Return",
]

# Master column letters mapped to customer columns (skip make, model, dates, internal fields)
CUSTOMER_MASTER_COLS = ["A", "E", "F", "C", "H", "I", "J", "K"]
MASTER_REF = f"'{MASTER_SHEET}'"


def style_range(ws, min_row, max_row, min_col, max_col, font=None, fill=None, border=None, alignment=None):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment


def copy_row_values(src_ws, dst_ws, src_row, dst_row, max_col=13):
    for c in range(1, max_col + 1):
        val = src_ws.cell(src_row, c).value
        if isinstance(val, str):
            val = val.strip()
        dst_ws.cell(dst_row, c, value=val)


def build_master(ws, src_ws):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "G.E.T. Rental Units - Master (Employees Only)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:O1")
    ws["A2"] = (
        "Enter and update all rental units here. "
        "The Customer Insurance sheet pulls from this table."
    )
    ws["A2"].font = SMALL_FONT
    ws.merge_cells("A2:O2")

    for idx, header in enumerate(MASTER_HEADERS, start=1):
        cell = ws.cell(DATA_START_ROW - 1, idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    last_src = src_ws.max_row
    for offset, src_row in enumerate(range(4, last_src + 1)):
        dst_row = DATA_START_ROW + offset
        if not any(src_ws.cell(src_row, c).value for c in range(1, 14)):
            continue
        copy_row_values(src_ws, ws, src_row, dst_row)

    last_data = ws.max_row
    for r in range(DATA_START_ROW, last_data + 1):
        stock = f"A{r}"
        vin = f"E{r}"
        date_in = f"M{r}"
        ws.cell(r, 14).value = f'=IF({stock}="","",IF({date_in}="","On Rent","Returned"))'
        ws.cell(r, 15).value = f'=IF({stock}="","",IF({vin}="","YES",""))'

    table_ref = f"A{DATA_START_ROW - 1}:O{max(last_data, DATA_START_ROW)}"
    tab = Table(displayName=TABLE_NAME, ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    widths = [12, 12, 8, 14, 22, 14, 10, 14, 14, 11, 11, 12, 12, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{DATA_START_ROW}"
    ws.row_dimensions[DATA_START_ROW - 1].height = 32

    fuel_vals = '"Full,3/4,1/2,1/4,Empty"'
    dv_fuel = DataValidation(type="list", formula1=f"={fuel_vals}", allow_blank=True)
    dv_fuel.error = "Choose a fuel level from the list."
    dv_fuel.errorTitle = "Invalid fuel level"
    ws.add_data_validation(dv_fuel)
    for col in ("J", "K"):
        dv_fuel.add(f"{col}{DATA_START_ROW}:{col}{DATA_START_ROW + MAX_DATA_ROWS}")

    dv_year = DataValidation(type="whole", operator="between", formula1="1980", formula2="2035", allow_blank=True)
    dv_year.errorTitle = "Invalid year"
    ws.add_data_validation(dv_year)
    dv_year.add(f"C{DATA_START_ROW}:C{DATA_START_ROW + MAX_DATA_ROWS}")

    vin_rule = FormulaRule(
        formula=[f'AND($A{DATA_START_ROW}<>"",$E{DATA_START_ROW}="")'],
        fill=WARN_FILL,
        font=Font(color=WARN_TEXT),
    )
    ws.conditional_formatting.add(
        f"A{DATA_START_ROW}:O{DATA_START_ROW + MAX_DATA_ROWS}",
        vin_rule,
    )

    style_range(
        ws,
        DATA_START_ROW,
        last_data,
        1,
        15,
        font=BODY_FONT,
        border=BORDER,
        alignment=Alignment(vertical="center"),
    )
    for r in range(DATA_START_ROW, last_data + 1):
        if (r - DATA_START_ROW) % 2 == 1:
            for c in range(1, 14):
                ws.cell(r, c).fill = ALT_FILL

    for r in range(DATA_START_ROW, last_data + 1):
        ws.cell(r, 1).number_format = "@"  # stock # (S1412 or numeric)
        ws.cell(r, 5).number_format = "@"  # VIN
        ws.cell(r, 6).number_format = "0"  # serial (no scientific notation)
        for c in (12, 13):
            ws.cell(r, c).number_format = "mm/dd/yyyy"
        ws.cell(r, 4).number_format = '"$"#,##0'
        # Normalize numeric stock IDs to text for consistency
        stock_val = ws.cell(r, 1).value
        if isinstance(stock_val, (int, float)) and stock_val == int(stock_val):
            ws.cell(r, 1, value=str(int(stock_val)))

    return last_data


def build_customer(ws, last_master_row: int):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "G.E.T. Equipment Rental - Insurance Reference"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        "Rental equipment details for your insurance carrier. "
        "G.E.T. maintains this list from our rental records."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:H2")

    header_row = 5
    for idx, header in enumerate(CUSTOMER_HEADERS, start=1):
        cell = ws.cell(header_row, idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    first_data = header_row + 1
    last_range_row = DATA_START_ROW + MAX_DATA_ROWS - 1
    current_rows = max(1, last_master_row - DATA_START_ROW + 1)
    formula_rows = min(MAX_DATA_ROWS, current_rows + 50)

    for offset in range(formula_rows):
        cust_row = first_data + offset
        index_expr = f"ROW()-{first_data}+1"
        for col_idx, master_col in enumerate(CUSTOMER_MASTER_COLS, start=1):
            col_range = (
                f"{MASTER_REF}!${master_col}${DATA_START_ROW}:"
                f"${master_col}${last_range_row}"
            )
            cell = ws.cell(
                cust_row,
                col_idx,
                value=(
                    f"=IF(INDEX({col_range},{index_expr})=\"\",\"\","
                    f"INDEX({col_range},{index_expr}))"
                ),
            )
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            if offset % 2 == 1:
                cell.fill = ALT_FILL
            if col_idx in (1, 2):
                cell.number_format = "@"
            elif col_idx == 3:
                cell.number_format = "0"

    widths = [14, 24, 14, 12, 16, 16, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{first_data}"
    ws.row_dimensions[header_row].height = 36

    ws.print_title_rows = f"${header_row}:${header_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def build_guide(ws):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "How to Use This Workbook"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "G.E.T. Rental Trucks"
    ws["A2"].font = SUBTITLE_FONT

    sections = [
        ("For G.E.T. employees", [
            "Open the Master - Rental Units sheet for all data entry.",
            "Add a new row inside the blue table (row below the last unit).",
            "Required for every unit: Stock #, Year, Serial Number, Meter Out, Fuel Out, Date Out.",
            "ON ROAD VIN is required for road-licensed units; rows missing VIN highlight in red.",
            "Rental Status and Needs VIN are formulas. Do not overwrite those cells.",
            "When a unit returns, fill Meter In, Fuel In, and Date In.",
            "Customer Insurance reads from the master table. Edit the master sheet only.",
        ]),
        ("Sharing with customers (Excel on the web)", [
            "Upload this file to OneDrive or SharePoint.",
            "In Excel for the web, use Share and set the link to view only.",
            "Before sharing, you can hide Master - Rental Units (right-click the sheet tab, Hide).",
            "Customers only need Customer Insurance and this page.",
            "PDF: open Customer Insurance, then File, Print, Save as PDF.",
        ]),
        ("For customers and insurance adjusters", [
            "Open the Customer Insurance sheet.",
            "Columns list unit ID, VIN, serial, model year, meters, and fuel levels.",
            "Make, model, and rental dates are on the master sheet only. Call G.E.T. if your carrier needs them.",
            "Attach this file or a PDF with your claim paperwork.",
        ]),
        ("Sheets in this file", [
            "How to Use (this page)",
            "Master - Rental Units (employee data entry)",
            "Customer Insurance (for insurance claims)",
        ]),
    ]

    row = 4
    for title, bullets in sections:
        ws.cell(row, 1, value=title).font = Font(name="Calibri", size=12, bold=True, color=NAVY)
        row += 1
        for bullet in bullets:
            ws.cell(row, 1, value=f"  - {bullet}").font = BODY_FONT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 95


def protect_sheets(wb):
    from openpyxl.styles import Protection

    master = wb[MASTER_SHEET]
    for r in range(DATA_START_ROW, DATA_START_ROW + MAX_DATA_ROWS + 50):
        for c in range(1, 14):
            master.cell(r, c).protection = Protection(locked=False)
        for c in (14, 15):
            master.cell(r, c).protection = Protection(locked=True)
    master.protection.sheet = True
    master.protection.formatCells = False
    master.protection.insertRows = True
    master.protection.deleteRows = True
    master.protection.autoFilter = True
    master.protection.sort = True

    customer = wb[CUSTOMER_SHEET]
    # Customer sheet stays unlocked so linked values display and copy cleanly


def main():
    if not SOURCE.exists():
        raise SystemExit(f"Source file not found: {SOURCE}")

    if not BACKUP.exists():
        shutil.copy2(SOURCE, BACKUP)

    src_wb = load_workbook(SOURCE)
    src_ws = src_wb.active

    wb = load_workbook(SOURCE)
    old_name = wb.sheetnames[0]
    master = wb[old_name]
    master.title = MASTER_SHEET

    # Clear old layout; rebuild from row 1
    for row in master.iter_rows():
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.border = Border()
    if master.tables:
        for t in list(master.tables.values()):
            del master.tables[t.name]

    last_master_row = build_master(master, src_ws)

    customer = wb.create_sheet(CUSTOMER_SHEET, 0)
    build_customer(customer, last_master_row)

    guide = wb.create_sheet(GUIDE_SHEET, 0)
    build_guide(guide)

    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True)
    wb.properties.title = "G.E.T. Rental Trucks"
    wb.properties.subject = "Rental trucks"
    wb.properties.creator = "G.E.T."
    wb.properties.description = "G.E.T. rental unit tracking"

    protect_sheets(wb)

    wb.active = guide

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")
    print(f"Backup: {BACKUP}")
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()

"""
Build technician daily payroll workbook with PTO and efficiency analytics.
Pay period: Wednesday through Tuesday (no weekend work). Paid weekly on Tuesday.
Excel 365 compatible (formulas, no VBA). Plain ranges (no tables) for reliable open/save.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.properties import CalcProperties

OUTPUT = Path(r"C:\Users\Jason Jones\Downloads\Technician-Daily-Payroll.xlsx")
WORKBOOK_AUTHOR = "Jason Jones"
# Dropdown list for pay period start dates (Pay Periods sheet, column A)
PAY_PERIOD_LIST_RANGE = "'Pay Periods'!$A$5:$A$500"

TECHS = [
    "Jaime Delaoch",
    "Shane Mccrary",
    "Kirk Magda",
    "Andy Stallings",
    "Judah Bullard",
    "John Harmon",
]

ADMINS = [
    "Ed Nelms",
    "Jason Jones",
    "Todd Chastain",
]

# Salaried admins: no clock in/out, just Hours Worked and PTO.
SALARIED_ADMINS = {"Ed Nelms", "Todd Chastain"}

ALL_EMPLOYEES = TECHS + ADMINS
# ~6 months of weekdays; smaller grid avoids Excel repair on open
MAX_LOG_ROWS = 130
PAY_PERIOD_START_RANGE = "'Pay Periods'!$A$5:$A$500"
ROSTER_EMP_RANGE = "Roster!$A$5:$A$20"
ROSTER_RATE_RANGE = "Roster!$C$5:$C$20"

DAILY_SHEET_REFS: dict[str, "DailySheetRefs"] = {}

FULL_TIME_WEEK_HRS = 40
UTIL_WEEK_CELL = "'Tech Weekly Hours'!$B$4"
WEEKLY_HOURS_SHEET = "Tech Weekly Hours"
STANDARD_LABOR_RATE = 108
JOHN_HARMON = "John Harmon"
JOHN_PNL_RATE = 34

TECH_HOURLY_PAY: dict[str, float] = {
    "Jaime Delaoch": 34,
    "Shane Mccrary": 34,
    "Kirk Magda": 28,
    "Andy Stallings": 34,
    "Judah Bullard": 21,
    "John Harmon": 34,
}

NAVY = "1F3864"
BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
LIGHT_GREEN = "E2EFDA"
LIGHT_YELLOW = "FFF2CC"
WHITE = "FFFFFF"
WARN = "FFC7CE"
WARN_TEXT = "9C0006"

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=10, color="404040")
BODY_FONT = Font(name="Calibri", size=11)
SMALL_FONT = Font(name="Calibri", size=9, color="666666")

HEADER_FILL = PatternFill("solid", fgColor=BLUE)
TITLE_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
CALC_FILL = PatternFill("solid", fgColor=LIGHT_YELLOW)
PTO_HEADER_FILL = PatternFill("solid", fgColor="548235")
WARN_FILL = PatternFill("solid", fgColor=WARN)
INPUT_FILL = PatternFill("solid", fgColor=LIGHT_YELLOW)

THIN = Side(style="thin", color="B4C6E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Pay period: Wed-Tue (5 work days: Wed-Fri, Mon-Tue). Payday = Tuesday (period end).
SHARED_HEADERS = ["Date", "Work Day", "Pay Period Start", "Pay Period End", "Payday"]
SUB_HEADERS = ["Clock In", "Lunch", "Clock Out", "PTO", "Billed", "Clocked"]
SUB_HEADER_LABELS = {
    "Clock In": "In",
    "Lunch": "Lunch",
    "Clock Out": "Out",
    "PTO": "PTO",
    "Billed": "Billed",
    "Clocked": "Clk Hrs",
}
SUB_HEADER_DISPLAY = {
    "Clock In": "Clock\nIn",
    "Lunch": "Lunch\n(min)",
    "Clock Out": "Clock\nOut",
    "PTO": "PTO",
    "Billed": "Billed",
    "Clocked": "Clocked\nHrs",
}
SHARED_HEADER_DISPLAY = {
    "Date": "Date",
    "Work Day": "Work\nDay",
    "Pay Period Start": "Period\nStart",
    "Pay Period End": "Period\nEnd",
    "Payday": "Payday",
}
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
WRAP_LEFT_INDENT = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

DASH_COL_WIDTHS = {"A": 28, "B": 14, "C": 12, "D": 22, "E": 14}

PAY_PERIOD_COL_WIDTHS = [26, 11, 13, 14, 12, 14, 16]
MONTHLY_COL_WIDTHS = [24, 10, 15, 12, 13, 15, 13, 18, 16]
PAY_PERIODS_COL_WIDTHS = [20, 20, 14, 16, 12, 28]
ROSTER_COL_WIDTHS = [26, 12, 18]


def _col_width_for_text(text: str, minimum: float = 10, maximum: float = 42) -> float:
    parts = str(text).split("\n")
    longest = max(len(p) for p in parts) if parts else minimum
    return min(maximum, max(minimum, longest * 1.15 + 3))


def _set_col_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _style_header_cells(ws, row: int, headers: list[str]) -> None:
    ws.row_dimensions[row].height = 42
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = WRAP_CENTER
        letter = get_column_letter(i)
        current = ws.column_dimensions[letter].width or 0
        need = _col_width_for_text(h)
        if current < need:
            ws.column_dimensions[letter].width = need


def _style_label_cell(ws, cell_ref: str, value: str) -> None:
    cell = ws[cell_ref]
    cell.value = value
    cell.font = Font(bold=True, color=NAVY)
    cell.alignment = WRAP_LEFT
    col = "".join(ch for ch in cell_ref if ch.isalpha())
    row = int("".join(ch for ch in cell_ref if ch.isdigit()))
    need = _col_width_for_text(value, minimum=14)
    if (ws.column_dimensions[col].width or 0) < need:
        ws.column_dimensions[col].width = need
    if (ws.row_dimensions[row].height or 15) < 30:
        ws.row_dimensions[row].height = 34


def _polish_sheet(
    ws,
    max_row: int | None = None,
    max_col: int | None = None,
    *,
    min_row: int = 1,
) -> None:
    """Word wrap and row heights for visible labels (skips formula cells for height)."""
    max_row = max_row or min(ws.max_row or 1, 420)
    max_col = max_col or min(ws.max_column or 1, 55)
    for r in range(min_row, max_row + 1):
        lines_in_row = 1
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            text = str(cell.value)
            is_formula = text.startswith("=")
            is_long_label = c == 1 and len(text) > 18 and not is_formula
            align = WRAP_LEFT if is_long_label else WRAP_CENTER
            cell.alignment = align
            if is_formula:
                continue
            col_letter = get_column_letter(c)
            col_w = ws.column_dimensions[col_letter].width or 12
            explicit_lines = text.count("\n") + 1
            wrapped_lines = max(
                explicit_lines,
                int(len(text.replace("\n", "")) / max(col_w * 1.15, 8)) + 1,
            )
            lines_in_row = max(lines_in_row, wrapped_lines)
        if r <= 2:
            ws.row_dimensions[r].height = max(30, lines_in_row * 16)
        elif lines_in_row > 1:
            ws.row_dimensions[r].height = max(22, min(lines_in_row * 15, 52))


SHARED_COL_WIDTHS_TECH = [12, 11, 13, 13, 12]
SHARED_COL_WIDTHS_ADMIN = [13, 12, 15, 15, 13]
SUB_COL_WIDTHS_TECH = [11, 9, 11, 9, 9, 10]
SUB_COL_WIDTHS_ADMIN = [13, 11, 13, 10, 10, 12]
COLS_PER_PERSON = len(SUB_HEADERS)
PERSON_FILLS = ["2E75B6", "548235", "7030A0", "C55A11", "00B0F0", "ED7D31"]

METRIC_TO_FIELD = {
    "Clocked (hrs)": "Clocked",
    "PTO (hrs)": "PTO",
}


def _short_name(full_name: str) -> str:
    return full_name.split()[0]


def _dashboard_name_col_width(people: list[str]) -> float:
    """Wide enough for full technician names in the weekly dashboard."""
    names = list(people) + ["ALL TECHNICIANS"]
    return max(_col_width_for_text(n, minimum=26, maximum=34) for n in names)


def _person_col(person: str, field: str) -> str:
    label = SUB_HEADER_LABELS.get(field, field)
    return f"{_short_name(person)} {label}"


@dataclass
class DailySheetRefs:
    sheet: str
    data_start: int
    data_end: int
    person_cols: dict[str, dict[str, str]]

    def col_rng(self, col: str) -> str:
        return f"'{self.sheet}'!${col}${self.data_start}:${col}${self.data_end}"

    def person_rng(self, person: str, field_key: str) -> str:
        return self.col_rng(self.person_cols[person][field_key])


def _daily_refs(sheet: str) -> DailySheetRefs:
    return DAILY_SHEET_REFS[sheet]


def _workday_crit(refs: DailySheetRefs) -> str:
    return f'{refs.col_rng("B")},"Workday"'


def _pay_period_crit(refs: DailySheetRefs, cell: str = "$B$4") -> str:
    return f"{_workday_crit(refs)},{refs.col_rng('C')},{cell}"


def _month_crit(refs: DailySheetRefs) -> str:
    return (
        f'{refs.col_rng("A")},">="&DATE($B$4,$D$4,1),'
        f'{refs.col_rng("A")},"<="&EOMONTH(DATE($B$4,$D$4,1),0),'
        f"{_workday_crit(refs)}"
    )


def _date_range_crit(refs: DailySheetRefs) -> str:
    return (
        f'{refs.col_rng("A")},">="&$H$4,{refs.col_rng("A")},"<="&$J$4,'
        f"{_workday_crit(refs)}"
    )


def _sumifs_person(refs: DailySheetRefs, person: str, metric: str, crit: str) -> str:
    field = METRIC_TO_FIELD[metric]
    return f"=SUMIFS({refs.person_rng(person, field)},{crit})"


def _count_days_person(refs: DailySheetRefs, person: str, crit: str) -> str:
    return f'=COUNTIFS({refs.person_rng(person, "Clock In")},"<>",{crit})'


def _sum_team_metric(refs: DailySheetRefs, people: list[str], metric: str, crit: str) -> str:
    field = METRIC_TO_FIELD[metric]
    parts = [f"SUMIFS({refs.person_rng(p, field)},{crit})" for p in people]
    return "=" + "+".join(parts)

PAY_PERIOD_HEADERS = [
    "Employee",
    "Role",
    "Days Logged",
    "Clocked (hrs)",
    "PTO (hrs)",
    "Overtime ($)",
    "Pay Cost ($)",
]

MONTHLY_HEADERS = [
    "Employee",
    "Role",
    "Pay Rate ($/hr)",
    "Days Logged",
    "Clocked (hrs)",
    "Pay Cost ($)",
    "Billed (hrs)",
    "Billed Revenue ($)",
    "$ Margin (P&L)",
]


def _roster_rate_lookup(row: int) -> str:
    return (
        f"IFERROR(INDEX({ROSTER_RATE_RANGE},"
        f"MATCH($A{row},{ROSTER_EMP_RANGE},0)),0)"
    )


def _pay_cost_ot_formula(refs: DailySheetRefs, person: str, crit: str, row: int) -> str:
    """Pay = clocked at rate (1.5x over 40) + PTO at straight rate."""
    clocked = _sumifs_person(refs, person, "Clocked (hrs)", crit).lstrip("=")
    pto = _sumifs_person(refs, person, "PTO (hrs)", crit).lstrip("=")
    rate = _roster_rate_lookup(row)
    reg = f"MIN({clocked},{FULL_TIME_WEEK_HRS})*{rate}"
    ot = f"MAX({clocked}-{FULL_TIME_WEEK_HRS},0)*{rate}*1.5"
    pto_pay = f"{pto}*{rate}"
    return f"={reg}+{ot}+{pto_pay}"


def _overtime_pay_formula(refs: DailySheetRefs, person: str, crit: str, row: int) -> str:
    """OT $ = clocked hours over 40 paid at 1.5x."""
    clocked = _sumifs_person(refs, person, "Clocked (hrs)", crit).lstrip("=")
    rate = _roster_rate_lookup(row)
    return f"=MAX({clocked}-{FULL_TIME_WEEK_HRS},0)*{rate}*1.5"


def _pay_cost_ot_monthly_formula(refs: DailySheetRefs, person: str, row: int) -> str:
    """Sum pay cost per pay week in month using multiplication (no IF in array)."""
    rate = _roster_rate_lookup(row)
    clocked_rng = refs.person_rng(person, "Clocked")
    pto_rng = refs.person_rng(person, "PTO")
    period = refs.col_rng("C")
    workday = f'{refs.col_rng("B")},"Workday"'
    week_clk = f"SUMIFS({clocked_rng},{period},{PAY_PERIOD_START_RANGE},{workday})"
    week_pto = f"SUMIFS({pto_rng},{period},{PAY_PERIOD_START_RANGE},{workday})"
    over = f"({week_clk}>{FULL_TIME_WEEK_HRS})*({week_clk}-{FULL_TIME_WEEK_HRS})"
    # reg hrs per week = clocked - over
    reg_pay = f"({week_clk}-{over})*{rate}"
    ot_pay = f"{over}*{rate}*1.5"
    pto_pay = f"{week_pto}*{rate}"
    return (
        "=SUMPRODUCT("
        f"--({PAY_PERIOD_START_RANGE}>=$G$4),"
        f"--({PAY_PERIOD_START_RANGE}<=$I$4),"
        f"({reg_pay})+({ot_pay})+({pto_pay}))"
    )


def _ot_monthly_pay_formula(refs: DailySheetRefs, person: str, row: int) -> str:
    """Overtime $ summed across pay weeks in month."""
    rate = _roster_rate_lookup(row)
    clocked_rng = refs.person_rng(person, "Clocked")
    period = refs.col_rng("C")
    workday = f'{refs.col_rng("B")},"Workday"'
    week_clk = f"SUMIFS({clocked_rng},{period},{PAY_PERIOD_START_RANGE},{workday})"
    over = f"({week_clk}>{FULL_TIME_WEEK_HRS})*({week_clk}-{FULL_TIME_WEEK_HRS})"
    ot_pay = f"{over}*{rate}*1.5"
    return (
        "=SUMPRODUCT("
        f"--({PAY_PERIOD_START_RANGE}>=$G$4),"
        f"--({PAY_PERIOD_START_RANGE}<=$I$4),"
        f"({ot_pay}))"
    )


def _pnl_labor_rate(person: str) -> str:
    if person == JOHN_HARMON:
        return str(JOHN_PNL_RATE)
    return str(STANDARD_LABOR_RATE)


def _sum_team_pay_cost(people: list[str], crit: str, row_start: int) -> str:
    parts = [f"F{row_start + i}" for i in range(len(people))]
    return "=" + "+".join(parts)


def _sum_col_range(letter: str, row_start: int, count: int) -> str:
    parts = [f"{letter}{row_start + i}" for i in range(count)]
    return "=" + "+".join(parts)

# Excel WEEKDAY(...,2): Mon=1 .. Sun=7. CHOOSE maps days back to Wednesday.
WORK_DAY_FORMULA = '=IF(A{row}="","",IF(OR(WEEKDAY(A{row},2)=6,WEEKDAY(A{row},2)=7),"Weekend","Workday"))'
PAY_PERIOD_START_FORMULA = '=IF(A{row}="","",A{row}-CHOOSE(WEEKDAY(A{row},2),5,6,0,1,2,3,4))'
PAY_PERIOD_END_FORMULA = '=IF(C{row}="","",C{row}+6)'
PAYDAY_FORMULA = '=IF(D{row}="","",D{row})'

# 735 displays "07:35", 1632 displays "16:32"; real time serials show hh:mm.
TIME_ENTRY_FORMAT = '[>=1]00":"00;hh:mm'


def _parse_time_cell(cell_ref: str) -> str:
    """07:15 serial, or HHMM 0715 / 1512."""
    return (
        f'IF(OR({cell_ref}="",{cell_ref}=0),"",'
        f'IF({cell_ref}<1,{cell_ref},'
        f'IF(AND({cell_ref}>=100,{cell_ref}<2400,MOD({cell_ref},100)<60),'
        f'TIME(INT({cell_ref}/100),MOD({cell_ref},100),0),"")))'
    )


def _lunch_minutes(cell_ref: str) -> str:
    """Minutes: 30, 60, or 1 (= one hour); time serial <1 = lunch duration."""
    return (
        f'IF({cell_ref}="",0,'
        f'IF({cell_ref}<1,ROUND({cell_ref}*1440,0),'
        f'IF({cell_ref}<=8,{cell_ref}*60,{cell_ref})))'
    )


def _auto_date_formula(
    row: int, data_start: int, clock_in_cols: list[str]
) -> str:
    """Next workday after row above when ANY person has clock-in on this row."""
    if row <= data_start:
        return ""
    prev = row - 1
    refs = ",".join(f"{c}{row}" for c in clock_in_cols)
    return (
        f'=IF(COUNTA({refs})=0,"",'
        f'IF(A{prev}="","",WORKDAY(A{prev},1)))'
    )


def _clocked_hours_formula(
    workday_col: str,
    row: int,
    clock_in: str,
    lunch: str,
    clock_out: str,
) -> str:
    t_in = _parse_time_cell(f"{clock_in}{row}")
    t_out = _parse_time_cell(f"{clock_out}{row}")
    lunch_min = _lunch_minutes(f"{lunch}{row}")
    return (
        f'=IF(${workday_col}{row}="Weekend",0,'
        f'IF(OR({t_in}="",{t_out}="",{t_out}<={t_in}),"",'
        f"MAX(0,({t_out}-{t_in})*24-{lunch_min}/60)))"
    )


def _add_time_entry_prompts(
    ws,
    data_start: int,
    data_end: int,
    person_blocks: list[tuple[str, int, int]],
) -> None:
    """Tooltip on clock columns: enter HHMM or h:mm."""
    dv = DataValidation(
        type="custom",
        formula1="TRUE",
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    for _person, start_col, _end_col in person_blocks:
        for offset in (0, 2):
            col = get_column_letter(start_col + offset)
            dv.add(f"{col}{data_start}:{col}{data_end}")


def _add_lunch_entry_prompts(
    ws,
    data_start: int,
    data_end: int,
    person_blocks: list[tuple[str, int, int]],
) -> None:
    dv = DataValidation(
        type="custom",
        formula1="TRUE",
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    for _person, start_col, _end_col in person_blocks:
        col = get_column_letter(start_col + 1)
        dv.add(f"{col}{data_start}:{col}{data_end}")


def _add_pay_period_dropdown(ws, cell: str) -> None:
    """List validation via sheet range (avoids Excel repair issues with table refs)."""
    dv = DataValidation(
        type="list",
        formula1=f"={PAY_PERIOD_LIST_RANGE}",
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(cell)


def _add_title_block(ws, title: str, subtitle: str, width: int = 13) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = WRAP_CENTER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = WRAP_CENTER
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = max(28, int(len(subtitle) / max(width, 1) * 4) + 18)


def _wednesday_on_or_before(d: date) -> date:
    # Monday=0 .. Sunday=6; Wednesday=2
    return d - timedelta(days=(d.weekday() - 2) % 7)


def _pay_period_rows(year: int) -> list[tuple[date, date, date]]:
    """Wed-Tue pay periods; payday on Tuesday (last day of period)."""
    start = _wednesday_on_or_before(date(year, 1, 1))
    if start.year < year:
        start += timedelta(days=7)
    periods: list[tuple[date, date, date]] = []
    while start <= date(year, 12, 31) + timedelta(days=7):
        end = start + timedelta(days=6)
        payday = end
        if start.year == year or end.year == year:
            periods.append((start, end, payday))
        start += timedelta(days=7)
    return periods


def _build_roster(wb: Workbook) -> None:
    ws = wb.create_sheet("Roster", 0)
    _add_title_block(
        ws,
        "Employee Roster",
        "Set hourly rates here. Techs vs Admin used on summary sheets.",
        4,
    )
    headers = ["Employee", "Role", "Hourly Rate ($)"]
    start = 4
    _style_header_cells(ws, start, headers)

    row = start + 1
    for name in TECHS:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value="Tech")
        ws.cell(row=row, column=3, value=TECH_HOURLY_PAY.get(name, ""))
        row += 1
    for name in ADMINS:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value="Admin")
        ws.cell(row=row, column=3, value="")
        row += 1

    end_row = row - 1
    _set_col_widths(ws, ROSTER_COL_WIDTHS)
    for r in range(5, end_row + 1):
        ws.row_dimensions[r].height = 22
        for c in range(1, 4):
            ws.cell(row=r, column=c).alignment = WRAP_CENTER
    _polish_sheet(ws, max_row=end_row + 2, max_col=3)


def _build_pay_periods(wb: Workbook) -> None:
    ws = wb.create_sheet("Pay Periods")
    _add_title_block(
        ws,
        "Pay Period Calendar",
        "Pay period: Wednesday through Tuesday. No weekend work. Paid weekly on Tuesday.",
        6,
    )
    headers = [
        "Pay Period Start",
        "Pay Period End",
        "Payday",
        "Work Days",
        "Calendar Year",
        "Week Label",
    ]
    start = 4
    _style_header_cells(ws, start, headers)

    row = start + 1
    years = (date.today().year - 1, date.today().year, date.today().year + 1)
    for year in years:
        for pstart, pend, payday in _pay_period_rows(year):
            ws.cell(row=row, column=1, value=pstart)
            ws.cell(row=row, column=2, value=pend)
            ws.cell(row=row, column=3, value=payday)
            ws.cell(row=row, column=4, value="Wed-Fri, Mon-Tue")
            ws.cell(row=row, column=5, value=year)
            ws.cell(row=row, column=6, value=f"{pstart.strftime('%b %d')} - {pend.strftime('%b %d')}")
            row += 1

    end_row = row - 1
    for c in (1, 2, 3):
        letter = get_column_letter(c)
        for r in range(start + 1, end_row + 1):
            ws[f"{letter}{r}"].number_format = "m/d/yyyy"
    _set_col_widths(ws, PAY_PERIODS_COL_WIDTHS)
    for r in range(5, end_row + 1):
        ws.row_dimensions[r].height = 20
    _polish_sheet(ws, max_row=end_row + 2, max_col=6)


def _tech_hours_for_week_formula(person: str, week_cell: str = UTIL_WEEK_CELL) -> str:
    refs = _daily_refs("Tech Daily")
    return (
        f"=SUMIFS({refs.person_rng(person, 'Clocked')},"
        f"{refs.col_rng('A')},\">=\"&{week_cell},"
        f"{refs.col_rng('A')},\"<=\"&{week_cell}+6,"
        f'{refs.col_rng("B")},"Workday")'
    )


def _build_tech_utilization_dashboard(ws, people: list[str]) -> int:
    """Weekly hours vs 40 hr baseline; returns last dashboard row."""
    title_row = 3
    filter_row = 4
    header_row = 5
    first_tech_row = 6
    last_tech_row = first_tech_row + len(people) - 1
    team_row = last_tech_row + 1

    ws.merge_cells(
        start_row=title_row, start_column=1, end_row=title_row, end_column=8,
    )
    title = ws.cell(row=title_row, column=1, value="WEEKLY HOURS - % OF 40 HR FULL-TIME WEEK")
    title.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.alignment = WRAP_CENTER
    ws.row_dimensions[title_row].height = 30

    _style_label_cell(ws, "A4", "Pay week start (Wed):")
    ws["B4"] = ""
    ws["B4"].number_format = "m/d/yyyy"
    ws["B4"].fill = INPUT_FILL
    ws["B4"].border = BORDER
    _add_pay_period_dropdown(ws, "B4")

    _style_label_cell(ws, "C4", "Week end:")
    ws["D4"] = '=IF(B4="","",B4+6)'
    ws["D4"].number_format = "m/d/yyyy"
    ws["D4"].fill = CALC_FILL
    ws["D4"].border = BORDER

    _style_label_cell(ws, "E4", "Full-time baseline:")
    ws["F4"] = FULL_TIME_WEEK_HRS
    ws["F4"].font = Font(bold=True, size=12, color=NAVY)
    ws["F4"].alignment = WRAP_CENTER
    _style_label_cell(ws, "G4", "hrs / tech")
    ws.row_dimensions[filter_row].height = 38

    dash_headers = ["Technician", "Hours Worked", "% of 40 Hrs", "Status", "Progress"]
    _style_header_cells(ws, header_row, dash_headers)

    for idx, person in enumerate(people):
        r = first_tech_row + idx
        ws.cell(row=r, column=1, value=person)
        ws.cell(row=r, column=1).fill = PatternFill(
            "solid", fgColor=PERSON_FILLS[idx % len(PERSON_FILLS)]
        )
        ws.cell(row=r, column=1).font = Font(bold=True, color=WHITE)
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = WRAP_LEFT_INDENT

        ws.cell(row=r, column=2, value=_tech_hours_for_week_formula(person))
        ws.cell(row=r, column=2).number_format = "0.00"
        ws.cell(row=r, column=2).fill = CALC_FILL
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=2).alignment = WRAP_CENTER

        ws.cell(row=r, column=3, value=f'=IF(B{r}="","",B{r}/{FULL_TIME_WEEK_HRS})')
        ws.cell(row=r, column=3).number_format = "0.0%"
        ws.cell(row=r, column=3).fill = CALC_FILL
        ws.cell(row=r, column=3).border = BORDER
        ws.cell(row=r, column=3).alignment = WRAP_CENTER

        ws.cell(
            row=r,
            column=4,
            value=(
                f'=IF(C{r}="","",IF(C{r}>=1,"Full week (100%+)",'
                f'IF(C{r}>=0.75,"On track",IF(C{r}>=0.5,"Partial week","Under 50%"))))'
            ),
        )
        ws.cell(row=r, column=4).font = Font(bold=True, size=10)
        ws.cell(row=r, column=4).border = BORDER
        ws.cell(row=r, column=4).alignment = WRAP_CENTER

        ws.cell(row=r, column=5, value=f'=IF(C{r}="","",C{r})')
        ws.cell(row=r, column=5).border = BORDER
        ws.row_dimensions[r].height = 26

    ws.cell(row=team_row, column=1, value="ALL TECHNICIANS").font = Font(
        bold=True, size=12, color=WHITE
    )
    ws.cell(row=team_row, column=1).fill = PatternFill("solid", fgColor="375623")
    ws.cell(row=team_row, column=1).border = BORDER
    ws.cell(row=team_row, column=1).alignment = WRAP_LEFT_INDENT

    ws.cell(row=team_row, column=2, value=f"=SUM(B{first_tech_row}:B{last_tech_row})")
    ws.cell(row=team_row, column=2).number_format = "0.00"
    ws.cell(row=team_row, column=2).font = Font(bold=True, size=12)
    ws.cell(row=team_row, column=2).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    ws.cell(row=team_row, column=2).border = BORDER

    cap = FULL_TIME_WEEK_HRS * len(people)
    ws.cell(row=team_row, column=3, value=f"=IF(B{team_row}=0,\"\",B{team_row}/{cap})")
    ws.cell(row=team_row, column=3).number_format = "0.0%"
    ws.cell(row=team_row, column=3).font = Font(bold=True, size=12)
    ws.cell(row=team_row, column=3).fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    ws.cell(row=team_row, column=3).border = BORDER

    ws.cell(
        row=team_row,
        column=4,
        value=f'="Total "&TEXT(B{team_row},"0.0")&" hrs ("&TEXT({cap},"0")&" hr capacity)"',
    )
    ws.cell(row=team_row, column=4).font = Font(bold=True, size=10, color="375623")
    ws.cell(row=team_row, column=4).border = BORDER
    ws.cell(row=team_row, column=4).alignment = WRAP_CENTER

    ws.cell(row=team_row, column=5, value=f'=IF(C{team_row}="","",C{team_row})')
    ws.row_dimensions[team_row].height = 28

    pct_range = f"C{first_tech_row}:C{team_row}"
    ws.conditional_formatting.add(
        pct_range,
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=0.75,
            mid_color="FFEB84",
            end_type="num",
            end_value=1,
            end_color="63BE7B",
        ),
    )
    ws.conditional_formatting.add(
        f"E{first_tech_row}:E{team_row}",
        DataBarRule(
            start_type="num",
            start_value=0,
            end_type="num",
            end_value=1,
            color=BLUE,
            showValue=False,
        ),
    )
    dash_widths = dict(DASH_COL_WIDTHS)
    dash_widths["A"] = _dashboard_name_col_width(people)
    for ref, width in dash_widths.items():
        ws.column_dimensions[ref].width = width

    return team_row


def _build_tech_weekly_hours_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(WEEKLY_HOURS_SHEET)
    _add_title_block(
        ws,
        "Weekly Hours - % of 40 HR Full-Time Week",
        "Pick pay week start (Wed). Hours pull from Tech Daily. Over 40 hrs = overtime (1.5x on pay sheets).",
        8,
    )
    _build_tech_utilization_dashboard(ws, TECHS)
    _fix_util_dashboard_layout(
        ws,
        TECHS,
        title_row=3,
        filter_row=4,
        header_row=5,
        first_tech_row=6,
        team_row=6 + len(TECHS),
    )
    _polish_sheet(ws, max_row=6 + len(TECHS) + 2, max_col=8)


def _fix_util_dashboard_layout(
    ws,
    people: list[str],
    *,
    title_row: int,
    filter_row: int,
    header_row: int,
    first_tech_row: int,
    team_row: int,
) -> None:
    """Keep dashboard rows compact and name column readable after sheet-wide polish."""
    widths = dict(DASH_COL_WIDTHS)
    widths["A"] = _dashboard_name_col_width(people)
    for ref, width in widths.items():
        ws.column_dimensions[ref].width = width

    ws.row_dimensions[title_row].height = 30
    ws.row_dimensions[filter_row].height = 34
    ws.row_dimensions[header_row].height = 42
    for r in range(first_tech_row, team_row):
        ws.row_dimensions[r].height = 26
    ws.row_dimensions[team_row].height = 28
    for r in range(first_tech_row, team_row + 1):
        ws.cell(row=r, column=1).alignment = WRAP_LEFT_INDENT
        for c in range(2, 6):
            ws.cell(row=r, column=c).alignment = WRAP_CENTER


def _team_daily_sheet(
    wb: Workbook,
    sheet_name: str,
    table_name: str,
    people: list[str],
    subtitle: str,
    *,
    is_admin: bool = False,
) -> None:
    """One row per day; each person has their own column block."""
    ws = wb.create_sheet(sheet_name)
    total_cols = len(SHARED_HEADERS) + len(people) * COLS_PER_PERSON
    last_col_letter = get_column_letter(total_cols)
    shared_widths = list(
        SHARED_COL_WIDTHS_ADMIN if is_admin else SHARED_COL_WIDTHS_TECH
    )
    sub_widths = SUB_COL_WIDTHS_ADMIN if is_admin else SUB_COL_WIDTHS_TECH
    zoom = 95 if is_admin else 80

    _add_title_block(ws, f"{sheet_name} - Daily Time & Payroll", subtitle, total_cols)

    hint_row = 3
    ws.merge_cells(start_row=hint_row, start_column=1, end_row=hint_row, end_column=total_cols)
    hint_cell = ws.cell(
        row=hint_row,
        column=1,
        value=(
            "One row per day. First date in column A; more dates auto-fill when you enter times. "
            "Clock In/Out: type 735 or 0735 (no colon needed; shows as 07:35). Lunch: 30 or 60 (minutes)."
        ),
    )
    hint_cell.font = SMALL_FONT
    hint_cell.alignment = WRAP_CENTER

    name_row = hint_row + 1
    label_row = hint_row + 2
    table_header_row = hint_row + 3
    data_start = hint_row + 4
    data_end = data_start + MAX_LOG_ROWS - 1
    workday_col = "B"

    for i, h in enumerate(SHARED_HEADERS, 1):
        for row, text in ((label_row, SHARED_HEADER_DISPLAY[h]), (table_header_row, h)):
            cell = ws.cell(row=row, column=i, value=text)
            cell.font = HEADER_FONT if row == table_header_row else Font(
                name="Calibri", size=9, bold=True, color=NAVY
            )
            fill = PatternFill("solid", fgColor="5B9BD5") if "Pay" in h or h == "Payday" else HEADER_FILL
            if h == "Work Day":
                fill = PatternFill("solid", fgColor="D9E1F2")
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = WRAP_CENTER

    person_blocks: list[tuple[str, int, int]] = []
    col_cursor = len(SHARED_HEADERS) + 1
    for pi, person in enumerate(people):
        start_col = col_cursor
        end_col = start_col + COLS_PER_PERSON - 1
        person_blocks.append((person, start_col, end_col))
        col_cursor = end_col + 1

        ws.merge_cells(
            start_row=name_row,
            start_column=start_col,
            end_row=name_row,
            end_column=end_col,
        )
        name_cell = ws.cell(row=name_row, column=start_col, value=person)
        name_cell.font = Font(name="Calibri", size=12 if is_admin else 11, bold=True, color=WHITE)
        name_cell.fill = PatternFill("solid", fgColor=PERSON_FILLS[pi % len(PERSON_FILLS)])
        name_cell.alignment = WRAP_CENTER
        name_cell.border = BORDER

        is_salaried = person in SALARIED_ADMINS

        for j, sub in enumerate(SUB_HEADERS):
            c = start_col + j
            disabled = is_salaried and sub in ("Clock In", "Lunch", "Clock Out", "Billed")
            sub_fill = PatternFill(
                "solid",
                fgColor=(
                    "D9D9D9"
                    if disabled
                    else (LIGHT_YELLOW if sub in ("PTO", "Billed", "Clocked") else LIGHT_BLUE)
                ),
            )
            if is_salaried and sub == "Clocked":
                label_text = "Hours\nWorked"
            elif disabled:
                label_text = "-"
            else:
                label_text = SUB_HEADER_DISPLAY[sub]
            label_cell = ws.cell(row=label_row, column=c, value=label_text)
            label_cell.font = Font(name="Calibri", size=9, bold=True, color=NAVY)
            label_cell.fill = sub_fill
            label_cell.border = BORDER
            label_cell.alignment = WRAP_CENTER

            hdr = ws.cell(row=table_header_row, column=c, value=_person_col(person, sub))
            hdr.font = Font(name="Calibri", size=9, bold=True, color=NAVY)
            hdr.fill = sub_fill
            hdr.border = BORDER
            hdr.alignment = WRAP_CENTER

    person_cols: dict[str, dict[str, str]] = {}
    for person, start_col, _end_col in person_blocks:
        person_cols[person] = {
            "Clock In": get_column_letter(start_col),
            "Lunch": get_column_letter(start_col + 1),
            "Clock Out": get_column_letter(start_col + 2),
            "PTO": get_column_letter(start_col + 3),
            "Billed": get_column_letter(start_col + 4),
            "Clocked": get_column_letter(start_col + 5),
        }
    DAILY_SHEET_REFS[sheet_name] = DailySheetRefs(
        sheet_name, data_start, data_end, person_cols
    )

    activity_cols: list[str] = []
    for person, start_col, _end_col in person_blocks:
        if person in SALARIED_ADMINS:
            activity_cols.append(get_column_letter(start_col + 5))
        else:
            activity_cols.append(get_column_letter(start_col))

    for r in range(data_start, data_end + 1):
        if r == data_start:
            ws.cell(row=r, column=1, value=None)
        else:
            ws.cell(row=r, column=1, value=_auto_date_formula(r, data_start, activity_cols))
        ws.cell(row=r, column=1).number_format = "m/d/yyyy"
        for c in (3, 4, 5):
            ws.cell(row=r, column=c).number_format = "m/d/yyyy"
        ws.cell(row=r, column=2, value=WORK_DAY_FORMULA.format(row=r))
        ws.cell(row=r, column=3, value=PAY_PERIOD_START_FORMULA.format(row=r))
        ws.cell(row=r, column=4, value=PAY_PERIOD_END_FORMULA.format(row=r))
        ws.cell(row=r, column=5, value=PAYDAY_FORMULA.format(row=r))
        for c in (2, 3, 4, 5):
            ws.cell(row=r, column=c).fill = CALC_FILL
            ws.cell(row=r, column=c).alignment = WRAP_CENTER

        ws.cell(row=r, column=1).alignment = WRAP_CENTER
        ws.cell(row=r, column=2).alignment = WRAP_CENTER

        for person, start_col, _end_col in person_blocks:
            is_salaried = person in SALARIED_ADMINS
            ci = get_column_letter(start_col)
            lunch = get_column_letter(start_col + 1)
            co = get_column_letter(start_col + 2)

            for c in range(start_col, start_col + COLS_PER_PERSON):
                ws.cell(row=r, column=c).alignment = WRAP_CENTER

            if is_salaried:
                gray_fill = PatternFill("solid", fgColor="F2F2F2")
                for c in (start_col, start_col + 1, start_col + 2, start_col + 4):
                    ws.cell(row=r, column=c).fill = gray_fill
                ws.cell(row=r, column=start_col + 3).number_format = "0.00"
                ws.cell(row=r, column=start_col + 5).number_format = "0.00"
                ws.cell(row=r, column=start_col + 5).fill = INPUT_FILL
            else:
                ws.cell(row=r, column=start_col).number_format = TIME_ENTRY_FORMAT
                ws.cell(row=r, column=start_col + 1).number_format = "0"
                ws.cell(row=r, column=start_col + 2).number_format = TIME_ENTRY_FORMAT
                for c in range(start_col + 3, start_col + 5):
                    ws.cell(row=r, column=c).number_format = "0.00"
                ws.cell(row=r, column=start_col + 5).number_format = "0.00"
                ws.cell(
                    row=r,
                    column=start_col + 5,
                    value=_clocked_hours_formula(workday_col, r, ci, lunch, co),
                )
                ws.cell(row=r, column=start_col + 5).fill = CALC_FILL

    for i, width in enumerate(shared_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for _person, start_col, _end_col in person_blocks:
        for j, width in enumerate(sub_widths):
            ws.column_dimensions[get_column_letter(start_col + j)].width = width
        # Billed column (offset 4) is no longer tracked; hide it but keep layout stable.
        billed_letter = get_column_letter(start_col + 4)
        ws.column_dimensions[billed_letter].hidden = True

    ws.freeze_panes = f"F{data_start}"
    ws.sheet_view.zoomScale = zoom
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[hint_row].height = 28
    ws.row_dimensions[name_row].height = 34 if is_admin else 30
    ws.row_dimensions[label_row].height = 48
    ws.row_dimensions[table_header_row].height = 22
    for r in range(data_start, data_end + 1):
        ws.row_dimensions[r].height = 22 if is_admin else 20
    ws.sheet_format.defaultRowHeight = 22 if is_admin else 20
    _polish_sheet(ws, min_row=1, max_row=2, max_col=min(total_cols, 12))
    _polish_sheet(
        ws,
        min_row=hint_row,
        max_row=data_end + 2,
        max_col=min(total_cols, 50),
    )
    ws.row_dimensions[hint_row].height = 28
    ws.row_dimensions[name_row].height = 34 if is_admin else 30
    ws.row_dimensions[label_row].height = 48
    ws.row_dimensions[table_header_row].height = 22


def _pay_period_rollup_block(ws, *, header_row: int) -> int:
    """Pay period hours and pay cost (no $ efficiency; see Monthly Summary)."""
    data_row_start = header_row + 1
    for idx, name in enumerate(ALL_EMPLOYEES):
        r = data_row_start + idx
        ws.cell(row=r, column=1, value=name)
        role = "Tech" if name in TECHS else "Admin"
        ws.cell(row=r, column=2, value=role)
        sheet = "Tech Daily" if role == "Tech" else "Admin Daily"
        refs = _daily_refs(sheet)
        crit = _pay_period_crit(refs)

        ws.cell(row=r, column=3, value=_count_days_person(refs, name, crit))
        ws.cell(row=r, column=4, value=_sumifs_person(refs, name, "Clocked (hrs)", crit))
        ws.cell(row=r, column=5, value=_sumifs_person(refs, name, "PTO (hrs)", crit))
        ws.cell(row=r, column=6, value=_overtime_pay_formula(refs, name, crit, r))
        ws.cell(row=r, column=7, value=_pay_cost_ot_formula(refs, name, crit, r))

        for c in range(1, 8):
            ws.cell(row=r, column=c).alignment = WRAP_CENTER
        for c in range(3, 8):
            ws.cell(row=r, column=c).fill = CALC_FILL
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=6).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=7).number_format = '"$"#,##0.00'
        for c in (4, 5):
            ws.cell(row=r, column=c).number_format = "0.00"

    return data_row_start + len(ALL_EMPLOYEES) - 1


def _build_pay_period_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Pay Period Summary")
    _add_title_block(
        ws,
        "Pay Period Payroll Summary",
        "Select pay period start (Wednesday). Period ends Tuesday; paid that Tuesday.",
        len(PAY_PERIOD_HEADERS),
    )
    _style_label_cell(ws, "A4", "Pay period start (Wednesday):")
    ws["B4"] = ""
    ws["B4"].number_format = "m/d/yyyy"
    ws["B4"].fill = INPUT_FILL
    ws["B4"].border = BORDER
    _style_label_cell(ws, "C4", "Pay period end (Tuesday):")
    ws["D4"] = '=IF(B4="","",B4+6)'
    ws["D4"].number_format = "m/d/yyyy"
    _style_label_cell(ws, "E4", "Payday (Tuesday):")
    ws["F4"] = '=IF(D4="","",D4)'
    ws["F4"].number_format = "m/d/yyyy"
    for ref in ("B4", "D4", "F4"):
        ws[ref].alignment = WRAP_CENTER
    for ref in ("D4", "F4"):
        ws[ref].fill = CALC_FILL
        ws[ref].border = BORDER
    ws.row_dimensions[4].height = 40

    _add_pay_period_dropdown(ws, "B4")

    header_row = 6
    _style_header_cells(ws, header_row, PAY_PERIOD_HEADERS)

    last = _pay_period_rollup_block(ws, header_row=header_row)

    sr = last + 3
    tech_start = header_row + 1
    ws.cell(row=sr, column=1, value="ALL TECHS (pay period)")
    ws.cell(row=sr, column=1).font = Font(bold=True, color=NAVY)
    ws.cell(row=sr, column=1).alignment = WRAP_LEFT
    ws.cell(
        row=sr,
        column=6,
        value="=" + "+".join(f"F{tech_start + i}" for i in range(len(TECHS))),
    )
    ws.cell(
        row=sr,
        column=7,
        value="=" + "+".join(f"G{tech_start + i}" for i in range(len(TECHS))),
    )
    for c in (6, 7):
        ws.cell(row=sr, column=c).number_format = '"$"#,##0.00'
        ws.cell(row=sr, column=c).fill = CALC_FILL
        ws.cell(row=sr, column=c).alignment = WRAP_CENTER
        ws.cell(row=sr, column=c).font = Font(bold=True)
    ws.row_dimensions[sr].height = 24

    _set_col_widths(ws, PAY_PERIOD_COL_WIDTHS)
    _polish_sheet(ws, max_row=sr + 2, max_col=7)


def _apply_pay_period_layout(ws, header_row: int = 6) -> None:
    _set_col_widths(ws, PAY_PERIOD_COL_WIDTHS)
    ws.row_dimensions[header_row].height = 42
    for r in range(header_row + 1, header_row + len(ALL_EMPLOYEES) + 1):
        ws.row_dimensions[r].height = 22


def _apply_summary_layout(ws, header_row: int = 6) -> None:
    _set_col_widths(ws, MONTHLY_COL_WIDTHS)
    ws.row_dimensions[header_row].height = 42
    for r in range(header_row + 1, header_row + len(TECHS) + 1):
        ws.row_dimensions[r].height = 22


def _build_monthly_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Monthly Summary")
    _add_title_block(
        ws,
        "Monthly $ Labor Efficiency (P&L)",
        f"Techs only. Type Billed (hrs) per tech in the yellow column. Billed Revenue = Billed hrs x ${STANDARD_LABOR_RATE}/hr (John ${JOHN_PNL_RATE}/hr). $ Margin = Billed Revenue - Pay Cost. Pay cost includes OT 1.5x over 40 hrs/week.",
        len(MONTHLY_HEADERS),
    )
    _style_label_cell(ws, "A4", "Year:")
    ws["B4"] = date.today().year
    ws["B4"].fill = INPUT_FILL
    ws["B4"].border = BORDER
    ws["B4"].alignment = WRAP_CENTER
    _style_label_cell(ws, "C4", "Month (1-12):")
    ws["D4"] = date.today().month
    ws["D4"].fill = INPUT_FILL
    ws["D4"].border = BORDER
    ws["D4"].alignment = WRAP_CENTER
    _style_label_cell(ws, "F4", "Month start:")
    ws["G4"] = "=DATE(B4,D4,1)"
    _style_label_cell(ws, "H4", "Month end:")
    ws["I4"] = "=EOMONTH(G4,0)"
    for ref in ("G4", "I4"):
        ws[ref].number_format = "m/d/yyyy"
        ws[ref].fill = CALC_FILL
        ws[ref].border = BORDER
        ws[ref].alignment = WRAP_CENTER
    ws.row_dimensions[4].height = 42

    dv_y = DataValidation(type="whole", operator="between", formula1="2024", formula2="2030", allow_blank=False)
    dv_m = DataValidation(type="whole", operator="between", formula1="1", formula2="12", allow_blank=False)
    ws.add_data_validation(dv_y)
    ws.add_data_validation(dv_m)
    dv_y.add("B4")
    dv_m.add("D4")

    header_row = 6
    _style_header_cells(ws, header_row, MONTHLY_HEADERS)

    data_row_start = header_row + 1
    tech_refs = _daily_refs("Tech Daily")

    for idx, name in enumerate(TECHS):
        r = data_row_start + idx
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value="Tech")
        crit = _month_crit(tech_refs)

        ws.cell(row=r, column=3, value=f"={_roster_rate_lookup(r)}")
        ws.cell(row=r, column=4, value=_count_days_person(tech_refs, name, crit))
        ws.cell(row=r, column=5, value=_sumifs_person(tech_refs, name, "Clocked (hrs)", crit))
        ws.cell(row=r, column=6, value=_pay_cost_ot_monthly_formula(tech_refs, name, r))
        # Billed (hrs) is a manual input per tech per month. Empty = 0.
        ws.cell(row=r, column=7, value=None)
        ws.cell(row=r, column=8, value=f'=IF(G{r}="",0,G{r})*{_pnl_labor_rate(name)}')
        ws.cell(row=r, column=9, value=f"=H{r}-F{r}")

        for c in range(1, 10):
            ws.cell(row=r, column=c).alignment = WRAP_CENTER
        for c in range(3, 10):
            ws.cell(row=r, column=c).border = BORDER
        for c in (3, 4, 5, 6, 8, 9):
            ws.cell(row=r, column=c).fill = CALC_FILL
        ws.cell(row=r, column=7).fill = INPUT_FILL
        ws.cell(row=r, column=3).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=6).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=7).number_format = "0.00"
        ws.cell(row=r, column=8).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=9).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=5).number_format = "0.00"

    last = data_row_start + len(TECHS) - 1
    tech_start = data_row_start
    sr = last + 3
    ws.cell(row=sr, column=1, value="ALL TECHS (monthly P&L)")
    ws.cell(row=sr, column=1).font = Font(bold=True, color=NAVY)
    ws.cell(row=sr, column=1).alignment = WRAP_LEFT
    ws.cell(row=sr, column=6, value=_sum_team_pay_cost(TECHS, "", tech_start))
    ws.cell(
        row=sr,
        column=7,
        value="=" + "+".join(f'IF(G{tech_start + i}="",0,G{tech_start + i})' for i in range(len(TECHS))),
    )
    ws.cell(row=sr, column=8, value="=" + "+".join(f"H{tech_start + i}" for i in range(len(TECHS))))
    ws.cell(row=sr, column=9, value=f"=H{sr}-F{sr}")
    for c in range(1, 10):
        ws.cell(row=sr, column=c).alignment = WRAP_CENTER
    for c in (6, 7, 8, 9):
        ws.cell(row=sr, column=c).fill = CALC_FILL
        ws.cell(row=sr, column=c).font = Font(bold=True)
    ws.cell(row=sr, column=6).number_format = '"$"#,##0.00'
    ws.cell(row=sr, column=7).number_format = "0.00"
    ws.cell(row=sr, column=8).number_format = '"$"#,##0.00'
    ws.cell(row=sr, column=9).number_format = '"$"#,##0.00'

    # Pay weeks with payday in selected month (for payroll accounting)
    pw_row = sr + 5
    ws.cell(row=pw_row, column=1, value="Paydays in this month (from Pay Periods)")
    ws.cell(row=pw_row, column=1).font = Font(bold=True, color=NAVY)
    ws.cell(row=pw_row, column=1).alignment = WRAP_LEFT
    ws.merge_cells(start_row=pw_row + 1, start_column=1, end_row=pw_row + 1, end_column=9)
    note = ws.cell(
        row=pw_row + 1,
        column=1,
        value=(
            '="Paydays in month: "&COUNTIFS(\'Pay Periods\'!$C$5:$C$500,">="&G4,'
            '\'Pay Periods\'!$C$5:$C$500,"<="&I4)&" (see Pay Periods sheet, column C)"'
        ),
    )
    note.alignment = WRAP_LEFT
    ws.row_dimensions[pw_row + 1].height = 28

    _set_col_widths(ws, MONTHLY_COL_WIDTHS)
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 16

    foot = ws.cell(
        row=pw_row + 3,
        column=1,
        value=(
            f"Type each tech's Billed (hrs) for the month in the yellow column G. "
            f"Billed Revenue = Billed hrs x ${STANDARD_LABOR_RATE}/hr (John Harmon at ${JOHN_PNL_RATE}/hr). "
            "$ Margin = Billed Revenue - Pay Cost. If Billed is empty Billed Revenue is $0, "
            "so $ Margin shows the negative tech expense (e.g. -$7,781.75). "
            "Pay cost uses OT (1.5x) for hours over 40 per pay week. Admins are not included on this sheet."
        ),
    )
    foot.font = SMALL_FONT
    foot.alignment = WRAP_LEFT
    ws.merge_cells(start_row=pw_row + 3, start_column=1, end_row=pw_row + 3, end_column=9)
    ws.row_dimensions[pw_row + 3].height = 46
    _polish_sheet(ws, max_row=pw_row + 4, max_col=9)


def _build_analytics(wb: Workbook) -> None:
    ws = wb.create_sheet("Analytics")
    _add_title_block(
        ws,
        "Hours & Pay Cost Report",
        "Custom date range: hours and pay cost only. Monthly $ P&L efficiency is on Monthly Summary.",
        len(PAY_PERIOD_HEADERS),
    )
    _style_label_cell(ws, "A4", "Pay period start (Wed):")
    ws["B4"] = ""
    ws["B4"].number_format = "m/d/yyyy"
    ws["B4"].fill = INPUT_FILL
    ws["B4"].border = BORDER
    ws["B4"].alignment = WRAP_CENTER
    _style_label_cell(ws, "C4", "Or report start:")
    ws["D4"] = ""
    ws["D4"].number_format = "m/d/yyyy"
    ws["D4"].fill = INPUT_FILL
    ws["D4"].border = BORDER
    ws["D4"].alignment = WRAP_CENTER
    _style_label_cell(ws, "E4", "Report end:")
    ws["F4"] = ""
    ws["F4"].number_format = "m/d/yyyy"
    ws["F4"].fill = INPUT_FILL
    ws["F4"].border = BORDER
    ws["F4"].alignment = WRAP_CENTER

    _add_pay_period_dropdown(ws, "B4")

    _style_label_cell(ws, "G4", "Effective start:")
    ws["H4"] = '=IF(B4<>"",B4,D4)'
    _style_label_cell(ws, "I4", "Effective end:")
    ws["J4"] = '=IF(B4<>"",B4+6,F4)'
    for ref in ("H4", "J4"):
        ws[ref].number_format = "m/d/yyyy"
        ws[ref].fill = CALC_FILL
        ws[ref].border = BORDER
        ws[ref].alignment = WRAP_CENTER
    ws.row_dimensions[4].height = 42

    header_row = 6
    _style_header_cells(ws, header_row, PAY_PERIOD_HEADERS)

    data_row_start = header_row + 1
    for idx, name in enumerate(TECHS):
        r = data_row_start + idx
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value="Tech")
        tech_refs = _daily_refs("Tech Daily")
        crit = _date_range_crit(tech_refs)

        ws.cell(row=r, column=3, value=_count_days_person(tech_refs, name, crit))
        ws.cell(row=r, column=4, value=_sumifs_person(tech_refs, name, "Clocked (hrs)", crit))
        ws.cell(row=r, column=5, value=_sumifs_person(tech_refs, name, "PTO (hrs)", crit))
        ws.cell(row=r, column=6, value=_overtime_pay_formula(tech_refs, name, crit, r))
        ws.cell(row=r, column=7, value=_pay_cost_ot_formula(tech_refs, name, crit, r))

        for c in range(1, 8):
            ws.cell(row=r, column=c).alignment = WRAP_CENTER
        for c in range(3, 8):
            ws.cell(row=r, column=c).fill = CALC_FILL
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=6).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=7).number_format = '"$"#,##0.00'
        for c in (4, 5):
            ws.cell(row=r, column=c).number_format = "0.00"

    last = data_row_start + len(TECHS) - 1
    sr = last + 3
    tech_start = data_row_start
    ws.cell(row=sr, column=1, value="ALL TECHS (date range)")
    ws.cell(row=sr, column=1).font = Font(bold=True, color=NAVY)
    ws.cell(row=sr, column=1).alignment = WRAP_LEFT
    ws.cell(
        row=sr, column=6,
        value="=" + "+".join(f"F{tech_start + i}" for i in range(len(TECHS))),
    )
    ws.cell(
        row=sr, column=7,
        value="=" + "+".join(f"G{tech_start + i}" for i in range(len(TECHS))),
    )
    for c in (6, 7):
        ws.cell(row=sr, column=c).alignment = WRAP_CENTER
        ws.cell(row=sr, column=c).fill = CALC_FILL
        ws.cell(row=sr, column=c).font = Font(bold=True)
    ws.cell(row=sr, column=6).number_format = '"$"#,##0.00'
    ws.cell(row=sr, column=7).number_format = '"$"#,##0.00'

    _set_col_widths(ws, PAY_PERIOD_COL_WIDTHS)
    _polish_sheet(ws, max_row=sr + 2, max_col=10)


def _build_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("How to Use")
    lines = [
        ("Technician Daily Payroll Workbook", TITLE_FONT),
        (f"File location: {OUTPUT}", SMALL_FONT),
        ("", BODY_FONT),
        ("Pay schedule", Font(bold=True, size=12, color=NAVY)),
        ("Pay period: Wednesday through Tuesday (paid weekly on Tuesday).", BODY_FONT),
        ("Work days only: Wednesday, Thursday, Friday, Monday, Tuesday. No Saturday or Sunday.", BODY_FONT),
        ("Tech Weekly Hours: % of 40 hr week per tech (pick pay week start on that sheet).", BODY_FONT),
        ("Tech Daily: 130 weekday rows (~6 months). First date in column A; more dates auto-fill.", BODY_FONT),
        ("Times as 07:15; lunch as 30 or 60 (minutes).", BODY_FONT),
        (
            "Clock times: 07:15 or 0715. Lunch: minutes (30, 60) or 1 for one hour. "
            "Weekly hours sum Mon-Fri rows in the pay week on Tech Weekly Hours.",
            BODY_FONT,
        ),
        ("Enter PTO under each name in the daily grid (Billed is no longer tracked daily).", BODY_FONT),
        ("", BODY_FONT),
        ("Pay Period Summary", Font(bold=True, size=12, color=NAVY)),
        ("Pick a pay period start (Wednesday) for weekly hours, overtime, and pay cost.", BODY_FONT),
        ("", BODY_FONT),
        ("Monthly Summary", Font(bold=True, size=12, color=NAVY)),
        ("Pick year and month. Hours sum by workday dates in that calendar month.", BODY_FONT),
        ("Type each tech's Billed Revenue for the month in the yellow column.", BODY_FONT),
        ("$ Margin (P&L) = Billed Revenue - Pay Cost. Empty Billed Revenue counts as $0, so $ Margin shows -Pay Cost.", BODY_FONT),
        ("Paydays in this month lists Tuesday pay dates in that month.", BODY_FONT),
        ("", BODY_FONT),
        ("Pay cost", Font(bold=True, size=12, color=NAVY)),
        ("Pay cost: first 40 hrs/week at roster rate; over 40 at 1.5x (time and a half). PTO paid at straight rate.", BODY_FONT),
        ("Monthly Summary and Analytics include technicians only (no admins).", BODY_FONT),
        ("Tech pay rates are on the Roster sheet.", BODY_FONT),
    ]
    for i, (text, font) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = font
        cell.alignment = WRAP_LEFT
        if text:
            ws.row_dimensions[i].height = max(20, min(72, 16 * (len(text) // 70 + 1)))
    ws.column_dimensions["A"].width = 96
    _polish_sheet(ws, max_row=len(lines) + 2, max_col=3)


def build() -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation = CalcProperties(fullCalcOnLoad=True, calcMode="auto")
    wb.properties.creator = WORKBOOK_AUTHOR
    wb.properties.lastModifiedBy = WORKBOOK_AUTHOR
    wb.properties.title = "Technician Daily Payroll"
    wb.properties.subject = "Technician daily time and payroll"
    wb.properties.description = "Daily time entry, pay periods, and payroll summaries."

    _build_roster(wb)
    _build_pay_periods(wb)
    _team_daily_sheet(
        wb,
        "Tech Daily",
        "TechDaily",
        TECHS,
        "Daily time entry only. See Tech Weekly Hours for the 40 hr week dashboard.",
        is_admin=False,
    )
    _build_tech_weekly_hours_sheet(wb)
    _team_daily_sheet(
        wb,
        "Admin Daily",
        "AdminDaily",
        ADMINS,
        "Admin staff - wider columns and larger text for easier entry.",
        is_admin=True,
    )
    _build_pay_period_summary(wb)
    _build_monthly_summary(wb)
    _build_analytics(wb)
    _build_instructions(wb)

    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")

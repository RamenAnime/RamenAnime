"""
Replace summary sheets (Pay Period Summary, Monthly Summary, Analytics) in the
existing payroll workbook. Hide Billed column on Tech Daily / Admin Daily.
Preserves every cell the user has typed on Tech Daily and Admin Daily.
"""
from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
BUILDER = SCRIPT_DIR / "build-tech-payroll-workbook.py"

spec = importlib.util.spec_from_file_location("payroll_builder", BUILDER)
assert spec and spec.loader
builder = importlib.util.module_from_spec(spec)
sys.modules["payroll_builder"] = builder
spec.loader.exec_module(builder)

OUTPUT = builder.OUTPUT
TECHS = builder.TECHS
DAILY_SHEET_REFS = builder.DAILY_SHEET_REFS


def _register_existing_refs(wb) -> None:
    """Rebuild in-memory refs from existing Tech Daily / Admin Daily."""
    from openpyxl.utils import get_column_letter

    for sheet_name, people in (("Tech Daily", builder.TECHS), ("Admin Daily", builder.ADMINS)):
        ws = wb[sheet_name]
        header_row = None
        for r in range(3, 16):
            if ws.cell(row=r, column=3).value == "Pay Period Start":
                header_row = r
                break
        if header_row is None:
            raise RuntimeError(f"Could not find header row in {sheet_name}")
        data_start = header_row + 1
        last = data_start
        for r in range(data_start, data_start + 500):
            if ws.cell(row=r, column=1).value is None and ws.cell(row=r, column=2).value is None:
                break
            last = r
        data_end = max(last, data_start + 129)

        person_cols: dict[str, dict[str, str]] = {}
        col_cursor = len(builder.SHARED_HEADERS) + 1
        for person in people:
            start_col = col_cursor
            person_cols[person] = {
                "Clock In": get_column_letter(start_col),
                "Lunch": get_column_letter(start_col + 1),
                "Clock Out": get_column_letter(start_col + 2),
                "PTO": get_column_letter(start_col + 3),
                "Billed": get_column_letter(start_col + 4),
                "Clocked": get_column_letter(start_col + 5),
            }
            col_cursor += 6  # existing daily sheets have 6 sub-cols per person
        DAILY_SHEET_REFS[sheet_name] = builder.DailySheetRefs(
            sheet_name, data_start, data_end, person_cols
        )


def _hide_billed_columns(wb) -> None:
    """Hide the Billed sub-column on Tech Daily and Admin Daily without touching data."""
    for sheet_name in ("Tech Daily", "Admin Daily"):
        if sheet_name not in wb.sheetnames:
            continue
        refs = DAILY_SHEET_REFS[sheet_name]
        for _person, cols in refs.person_cols.items():
            billed_col = cols["Billed"]
            wb[sheet_name].column_dimensions[billed_col].hidden = True


def main() -> None:
    if not OUTPUT.exists():
        print(f"Workbook not found at {OUTPUT}; nothing to fix.")
        return

    wb = load_workbook(OUTPUT)
    _register_existing_refs(wb)
    _hide_billed_columns(wb)

    sheets_to_replace = ["Pay Period Summary", "Monthly Summary", "Analytics"]
    for name in sheets_to_replace:
        if name in wb.sheetnames:
            del wb[name]

    builder._build_pay_period_summary(wb)
    builder._build_monthly_summary(wb)
    builder._build_analytics(wb)

    desired_order = [
        "Roster",
        "Pay Periods",
        "Tech Daily",
        "Tech Weekly Hours",
        "Admin Daily",
        "Pay Period Summary",
        "Monthly Summary",
        "Analytics",
        "How to Use",
    ]
    sheets_ordered = [n for n in desired_order if n in wb.sheetnames]
    others = [n for n in wb.sheetnames if n not in sheets_ordered]
    wb._sheets = [wb[n] for n in sheets_ordered + others]

    wb.save(OUTPUT)
    print(f"Updated summary sheets in {OUTPUT}")
    print("Billed column hidden on Tech Daily and Admin Daily.")
    print("Tech Daily and Admin Daily data preserved.")


if __name__ == "__main__":
    main()
